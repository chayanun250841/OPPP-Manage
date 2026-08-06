"""Per-facility service reconciliation pipeline (admin only -- has PID/name).

Follows this exact workflow:
  1. Group raw records by HCODE
  2. Each HCODE's people: PID, name, PP, FS
  3. Per record, predict service(s) from PP and FS separately using combo
     matching against the 15-item claim-rate list (rate_claim = full SPSC
     schedule rate, which is what actually appears in the raw PP/FS amounts)
  4. Summarize predicted item counts for the facility
  5. Reconciliation: value at SPSC's full rate (rate_claim) vs value actually
     allocated to the facility per the provincial agreement
     (rate_facility_share)

Prediction runs on PP and FS separately (not combined into ยอดรวม) because the
15 reference items are PP Fee schedule items -- mixing in FS would corrupt the
match. Each record's PP and FS amounts are decomposed independently.
"""
from __future__ import annotations

import json
import os
from html import escape

import pandas as pd

import db
import service_matching

_RATES_PATH = os.path.join(os.path.dirname(__file__), "assets", "service_rates.json")


def _load_items() -> list[dict]:
    try:
        with open(_RATES_PATH, "r", encoding="utf-8") as file:
            data = json.load(file)
    except (OSError, json.JSONDecodeError):
        return []
    return [item for item in data.get("items", []) if item.get("matchable")]


MATCHABLE_ITEMS = _load_items()
ITEM_NAMES = [item["name"] for item in MATCHABLE_ITEMS]
RATE_FULL_BY_NAME = {item["name"]: float(item["rate_claim"]) for item in MATCHABLE_ITEMS}
RATE_PROVINCIAL_BY_NAME = {
    item["name"]: (float(item["rate_facility_share"]) if item.get("rate_facility_share") is not None else None)
    for item in MATCHABLE_ITEMS
}

PEOPLE_COLUMNS = ["HCODE", "PID", "ชื่อ-นามสกุล", "PP", "FS", "ยอดรวม"]
PREDICTION_COLUMNS = ["PID", "ชื่อ-นามสกุล", "PP", "FS", "บริการที่คาดการณ์ (PP)", "บริการที่คาดการณ์ (FS)", "สถานะ"]
COUNT_COLUMNS = ["รายการบริการ", "จำนวนครั้ง"]
RECONCILE_COLUMNS = [
    "รายการบริการ", "จำนวนครั้ง",
    "อัตราเต็ม สปสช. (บาท/ครั้ง)", "ยอดที่สปสช.ชดเชย (บาท)",
    "อัตราจังหวัด (บาท/ครั้ง)", "ยอดที่ได้รับจัดสรรจริง (บาท)",
]

_STATUS_RANK = {"🔴 ไม่พบ": 0, "🟠 ใกล้เคียง": 1, "🟡 ไม่แน่ชัด": 2, "🟢 คาดการณ์": 3, "-": 4}


def get_people_for_hcode(hcode: str) -> pd.DataFrame:
    """Step 1+2: raw people list for one facility."""
    try:
        people = db.get_people_records_for_hcode(hcode)
    except Exception:  # noqa: BLE001
        return pd.DataFrame(columns=PEOPLE_COLUMNS)
    if people.empty:
        return pd.DataFrame(columns=PEOPLE_COLUMNS)
    return people[PEOPLE_COLUMNS]


def predict_records(people: pd.DataFrame) -> pd.DataFrame:
    """Step 3: per-record combo prediction, PP and FS decomposed separately."""
    if people.empty:
        return pd.DataFrame(columns=PREDICTION_COLUMNS)

    rows = []
    for rec in people.to_dict(orient="records"):
        pp_status, pp_combos, pp_diff = service_matching.resolve_combo(float(rec["PP"]))
        fs_status, fs_combos, fs_diff = service_matching.resolve_combo(float(rec["FS"]))
        overall_status = min(pp_status, fs_status, key=lambda s: _STATUS_RANK[s])
        rows.append({
            "PID": rec["PID"],
            "ชื่อ-นามสกุล": rec["ชื่อ-นามสกุล"],
            "PP": rec["PP"],
            "FS": rec["FS"],
            "บริการที่คาดการณ์ (PP)": service_matching.format_combo_text(pp_status, pp_combos, pp_diff),
            "บริการที่คาดการณ์ (FS)": service_matching.format_combo_text(fs_status, fs_combos, fs_diff),
            "สถานะ": overall_status,
            "_pp_status": pp_status,
            "_pp_combo": pp_combos[0] if pp_status in ("🟢 คาดการณ์", "🟠 ใกล้เคียง") else [],
            "_pp_diff": pp_diff,
            "_fs_status": fs_status,
            "_fs_combo": fs_combos[0] if fs_status in ("🟢 คาดการณ์", "🟠 ใกล้เคียง") else [],
            "_fs_diff": fs_diff,
        })
    return pd.DataFrame(rows)


def summarize_item_counts(predictions: pd.DataFrame) -> pd.DataFrame:
    """Step 4: count item hits across the facility. Confident (🟢) and
    closest-match (🟠) hits are both attributed to their matched item, since
    the goal is to match amounts as closely as possible; truly ambiguous (🟡)
    and unmatched (🔴) amounts are kept visible as their own catch-all rows
    rather than silently dropped or guessed into a specific item."""
    if predictions.empty:
        return pd.DataFrame(columns=COUNT_COLUMNS)

    counts: dict[str, int] = {name: 0 for name in ITEM_NAMES}
    unclear = 0
    notfound = 0
    approx_count = 0
    approx_diff_total = 0.0
    for rec in predictions.to_dict(orient="records"):
        for status_key, combo_key, diff_key in (
            ("_pp_status", "_pp_combo", "_pp_diff"),
            ("_fs_status", "_fs_combo", "_fs_diff"),
        ):
            status = rec[status_key]
            if status in ("🟢 คาดการณ์", "🟠 ใกล้เคียง"):
                for name in rec[combo_key]:
                    counts[name] = counts.get(name, 0) + 1
                if status == "🟠 ใกล้เคียง":
                    approx_count += 1
                    approx_diff_total += rec[diff_key]
            elif status == "🟡 ไม่แน่ชัด":
                unclear += 1
            elif status == "🔴 ไม่พบ":
                notfound += 1

    rows = [{"รายการบริการ": name, "จำนวนครั้ง": count} for name, count in counts.items() if count > 0]
    if approx_count:
        rows.append({
            "รายการบริการ": f"🟠 นับรวมข้างต้นแล้ว {approx_count} รายการเป็นการจับคู่แบบใกล้เคียง (ส่วนต่างรวม {approx_diff_total:,.2f} บาท)",
            "จำนวนครั้ง": approx_count,
        })
    if unclear:
        rows.append({"รายการบริการ": "🟡 ยังไม่แน่ชัด (ต้องตรวจสอบ)", "จำนวนครั้ง": unclear})
    if notfound:
        rows.append({"รายการบริการ": "🔴 ไม่พบรายการที่ตรงกัน", "จำนวนครั้ง": notfound})
    if not rows:
        return pd.DataFrame(columns=COUNT_COLUMNS)
    return pd.DataFrame(rows).sort_values("จำนวนครั้ง", ascending=False).reset_index(drop=True)


def build_reconciliation(counts: pd.DataFrame) -> pd.DataFrame:
    """Step 5: SPSC full-rate value vs actual provincial-agreement value."""
    if counts.empty:
        return pd.DataFrame(columns=RECONCILE_COLUMNS)

    rows = []
    total_full = 0.0
    total_provincial = 0.0
    for rec in counts.to_dict(orient="records"):
        name = rec["รายการบริการ"]
        count = int(rec["จำนวนครั้ง"])
        rate_full = RATE_FULL_BY_NAME.get(name)
        rate_provincial = RATE_PROVINCIAL_BY_NAME.get(name)
        if rate_full is None:
            rows.append({
                "รายการบริการ": name, "จำนวนครั้ง": f"{count:,}",
                "อัตราเต็ม สปสช. (บาท/ครั้ง)": "-", "ยอดที่สปสช.ชดเชย (บาท)": "-",
                "อัตราจังหวัด (บาท/ครั้ง)": "-", "ยอดที่ได้รับจัดสรรจริง (บาท)": "-",
            })
            continue
        full_value = count * rate_full
        total_full += full_value
        if rate_provincial is None:
            rows.append({
                "รายการบริการ": name, "จำนวนครั้ง": f"{count:,}",
                "อัตราเต็ม สปสช. (บาท/ครั้ง)": f"{rate_full:,.2f}",
                "ยอดที่สปสช.ชดเชย (บาท)": f"{full_value:,.2f}",
                "อัตราจังหวัด (บาท/ครั้ง)": "ยังไม่ยืนยัน",
                "ยอดที่ได้รับจัดสรรจริง (บาท)": "ยังไม่ยืนยัน",
            })
            continue
        provincial_value = count * rate_provincial
        total_provincial += provincial_value
        rows.append({
            "รายการบริการ": name, "จำนวนครั้ง": f"{count:,}",
            "อัตราเต็ม สปสช. (บาท/ครั้ง)": f"{rate_full:,.2f}",
            "ยอดที่สปสช.ชดเชย (บาท)": f"{full_value:,.2f}",
            "อัตราจังหวัด (บาท/ครั้ง)": f"{rate_provincial:,.2f}",
            "ยอดที่ได้รับจัดสรรจริง (บาท)": f"{provincial_value:,.2f}",
        })

    rows.append({
        "รายการบริการ": "รวม (เฉพาะรายการที่ยืนยันชัดเจน)", "จำนวนครั้ง": "-",
        "อัตราเต็ม สปสช. (บาท/ครั้ง)": "-", "ยอดที่สปสช.ชดเชย (บาท)": f"{total_full:,.2f}",
        "อัตราจังหวัด (บาท/ครั้ง)": "-", "ยอดที่ได้รับจัดสรรจริง (บาท)": f"{total_provincial:,.2f}",
    })
    return pd.DataFrame(rows, columns=RECONCILE_COLUMNS)


def format_people_display(people: pd.DataFrame) -> pd.DataFrame:
    if people.empty:
        return people
    display = people.copy()
    for col in ("PP", "FS", "ยอดรวม"):
        display[col] = display[col].map(lambda v: f"{float(v):,.2f}")
    return display


def format_predictions_display(predictions: pd.DataFrame) -> pd.DataFrame:
    if predictions.empty:
        return pd.DataFrame(columns=PREDICTION_COLUMNS)
    display = predictions[PREDICTION_COLUMNS].copy()
    display["PP"] = display["PP"].map(lambda v: f"{float(v):,.2f}")
    display["FS"] = display["FS"].map(lambda v: f"{float(v):,.2f}")
    return display


def analyze_hcode(hcode: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Run the full 5-step pipeline for one facility."""
    people = get_people_for_hcode(hcode)
    predictions = predict_records(people)
    counts = summarize_item_counts(predictions)
    reconciliation = build_reconciliation(counts)
    return (
        format_people_display(people),
        format_predictions_display(predictions),
        counts,
        reconciliation,
    )


# ---------------------------------------------------------------------------
# All-facilities pivot: the final destination view, one row per facility,
# replacing the manual "PP free Schedule" spreadsheet staff fill out today.
# ---------------------------------------------------------------------------

TOTAL_ROW_LABEL = "รวมทั้งหมด"

# Every service contributes the same pair of sub-columns, so the pivot reads as
# one grouped header ("service" spanning "จำนวนครั้ง" + "ยอดชดเชย") rather than a
# long run of differently-worded labels. Column keys join the two with " | ".
COUNT_SUB = "จำนวนครั้ง"
AMOUNT_SUB = "ยอดชดเชย"
LABEL_COLUMNS = ("HCODE", "ชื่อหน่วยบริการ")


def item_header(item: dict) -> str:
    """Column label for the pivot. Uses the item's short name -- a bare code
    like '2.1)' told the reader nothing, and the full name is far too long for
    a column that only holds a count."""
    return item.get("short") or item["name"]


def build_item_legend() -> str:
    """Short name -> full official name, for the readers who need the exact
    wording behind the shortened column headers."""
    lines = [
        f"**{item_header(item)}** = {item['name']} ({item['rate_claim']:,.0f} บาท · รหัส {item['code']})"
        for item in MATCHABLE_ITEMS
    ]
    return "\n\n".join(lines)


def build_all_facilities_pivot(hcode_names: dict[str, str]) -> pd.DataFrame:
    try:
        totals_by_hcode = db.get_summary_by_hcode()
    except Exception:  # noqa: BLE001
        return pd.DataFrame()
    if totals_by_hcode.empty:
        return pd.DataFrame()

    rows = []
    for rec in totals_by_hcode.to_dict(orient="records"):
        hcode = str(rec["HCODE"])
        facility_total = float(rec["ยอดรวม"])

        people = get_people_for_hcode(hcode)
        predictions = predict_records(people)
        counts = summarize_item_counts(predictions)
        count_by_item = {c["รายการบริการ"]: int(c["จำนวนครั้ง"]) for c in counts.to_dict(orient="records")}

        row: dict[str, object] = {"HCODE": hcode, "ชื่อหน่วยบริการ": hcode_names.get(hcode, "")}
        matched_claim_total = 0.0
        matched_share_total = 0.0
        for item in MATCHABLE_ITEMS:
            name = item["name"]
            label = item_header(item)
            count = count_by_item.get(name, 0)
            claim_amount = count * item["rate_claim"]
            share_rate = item.get("rate_facility_share")
            share_amount = count * share_rate if share_rate is not None else 0.0
            row[f"{label} | {COUNT_SUB}"] = count
            row[f"{label} | {AMOUNT_SUB}"] = claim_amount
            matched_claim_total += claim_amount
            matched_share_total += share_amount

        row[f"ยังไม่จัดประเภท | {AMOUNT_SUB}"] = round(facility_total - matched_claim_total, 2)
        row[f"รวมจาก สปสช. | {AMOUNT_SUB}"] = round(facility_total, 2)
        row[f"จัดสรรตามมติจังหวัด | {AMOUNT_SUB}"] = round(matched_share_total, 2)
        rows.append(row)

    pivot = pd.DataFrame(rows)
    numeric_cols = [c for c in pivot.columns if c not in ("HCODE", "ชื่อหน่วยบริการ")]
    total_row: dict[str, object] = {"HCODE": TOTAL_ROW_LABEL, "ชื่อหน่วยบริการ": ""}
    for col in numeric_cols:
        total_row[col] = pivot[col].sum()
    pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)
    return pivot


def format_all_facilities_pivot(pivot: pd.DataFrame) -> pd.DataFrame:
    """Flatten the grouped columns into one label per column. Kept for any
    caller that wants a plain frame; the Excel export uses the merged-header
    writer below so the workbook matches what the page shows."""
    if pivot.empty:
        return pivot
    display = pivot.copy()
    for col in display.columns:
        if col in LABEL_COLUMNS:
            continue
        if col.endswith(COUNT_SUB):
            display[col] = display[col].map(lambda v: f"{int(v):,}")
        else:
            display[col] = display[col].map(lambda v: f"{float(v):,.2f}")
    display.columns = [c.replace(" | ", " - ") for c in display.columns]
    return display


def _column_groups(value_columns: list[str]) -> list[tuple[str, list[str]]]:
    groups: list[tuple[str, list[str]]] = []
    for column in value_columns:
        group, sub = _split_column(column)
        if groups and groups[-1][0] == group:
            groups[-1][1].append(sub)
        else:
            groups.append((group, [sub]))
    return groups


def write_all_facilities_excel(pivot: pd.DataFrame, path: str) -> None:
    """Write the pivot with the same two-row merged header the page shows.

    pandas' to_excel would flatten the grouped columns into one label per
    column, so the workbook stopped looking like the table it came from.
    Values are written as real numbers with display formats, so totals and
    sorting still work in the spreadsheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    value_columns = [c for c in pivot.columns if c not in LABEL_COLUMNS]
    groups = _column_groups(value_columns)

    book = Workbook()
    sheet = book.active
    sheet.title = "สรุปทุกหน่วยบริการ"

    thin = Side(style="thin", color="B7C9BD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="E8F1EB")
    head_font = Font(bold=True, color="0F5C33")
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for index, label in enumerate(LABEL_COLUMNS, start=1):
        sheet.cell(row=1, column=index, value=label)
        sheet.merge_cells(start_row=1, start_column=index, end_row=2, end_column=index)

    column = len(LABEL_COLUMNS) + 1
    for group, subs in groups:
        sheet.cell(row=1, column=column, value=group)
        if len(subs) > 1:
            sheet.merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + len(subs) - 1)
        for offset, sub in enumerate(subs):
            sheet.cell(row=2, column=column + offset, value=sub)
        column += len(subs)

    for row in sheet["A1":f"{get_column_letter(column - 1)}2"]:
        for cell in row:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = centre
            cell.border = border

    for offset, record in enumerate(pivot.to_dict(orient="records")):
        row_index = 3 + offset
        is_total = str(record.get("HCODE", "")) == TOTAL_ROW_LABEL
        for index, label in enumerate(LABEL_COLUMNS, start=1):
            cell = sheet.cell(row=row_index, column=index, value=record.get(label, ""))
            cell.border = border
            if is_total:
                cell.font = Font(bold=True)
        for index, name in enumerate(value_columns, start=len(LABEL_COLUMNS) + 1):
            _group, sub = _split_column(name)
            raw = record.get(name, 0)
            value = int(raw) if sub == COUNT_SUB else round(float(raw), 2)
            cell = sheet.cell(row=row_index, column=index, value=value)
            cell.number_format = "#,##0" if sub == COUNT_SUB else "#,##0.00"
            cell.border = border
            if is_total:
                cell.font = Font(bold=True)

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 38
    for index in range(len(LABEL_COLUMNS) + 1, column):
        sheet.column_dimensions[get_column_letter(index)].width = 14
    sheet.freeze_panes = "C3"
    book.save(path)


def _split_column(column: str) -> tuple[str, str]:
    group, _, sub = column.partition(" | ")
    return group, sub


def render_all_facilities_html(pivot: pd.DataFrame) -> str:
    """Render the pivot with a two-row header: each service name spans its
    จำนวนครั้ง and ยอดชดเชย columns.

    gr.Dataframe cannot merge header cells, so this table is emitted as HTML.
    The numbers are the same ones format_all_facilities_pivot() exports.
    """
    if pivot.empty:
        return "<p class='hint-text'>ยังไม่มีข้อมูล — กดปุ่มคำนวณหลังอัปโหลดไฟล์แล้ว</p>"

    value_columns = [c for c in pivot.columns if c not in LABEL_COLUMNS]

    # Consecutive columns sharing a group become one spanning header cell.
    groups: list[tuple[str, list[str]]] = []
    for column in value_columns:
        group, sub = _split_column(column)
        if groups and groups[-1][0] == group:
            groups[-1][1].append(sub)
        else:
            groups.append((group, [sub]))

    head = ["<thead><tr>"]
    for label in LABEL_COLUMNS:
        head.append(f'<th class="lbl" rowspan="2">{escape(label)}</th>')
    for group, subs in groups:
        head.append(f'<th class="grp" colspan="{len(subs)}">{escape(group)}</th>')
    head.append("</tr><tr>")
    for _group, subs in groups:
        for sub in subs:
            head.append(f'<th class="sub">{escape(sub)}</th>')
    head.append("</tr></thead>")

    body = ["<tbody>"]
    for record in pivot.to_dict(orient="records"):
        is_total = str(record.get("HCODE", "")) == TOTAL_ROW_LABEL
        body.append(f'<tr class="{"total" if is_total else ""}">')
        for label in LABEL_COLUMNS:
            body.append(f'<td class="lbl">{escape(str(record.get(label, "")))}</td>')
        for column in value_columns:
            value = record.get(column, 0)
            _group, sub = _split_column(column)
            text = f"{int(value):,}" if sub == COUNT_SUB else f"{float(value):,.2f}"
            body.append(f"<td>{text}</td>")
        body.append("</tr>")
    body.append("</tbody>")

    return f'<div class="pivot-scroll"><table class="pivot">{"".join(head)}{"".join(body)}</table></div>'
